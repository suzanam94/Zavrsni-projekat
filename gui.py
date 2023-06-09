from knjizara import *
import tkinter as tk
from PIL import ImageTk, Image
from tkinter import messagebox
import openpyxl as openpyxl
import pyautogui as pyautogui
import matplotlib.pyplot as plt
import json as json

def kupac_prozor():
    kupac_prozor = tk.Toplevel(root)
    kupac_prozor.title("Kupac")
    proizvodi_u_korpi = []

    def ucitaj_proizvode():
        rezultat = proizvodi.ucitaj_proizvode()
        return rezultat
    
    def dodaj_u_korpu():
        selected_product = proizvod_listbox.get(tk.ACTIVE)
        korpa_listbox.insert(tk.END, selected_product)

        # Dodaj izabrani proizvod u listu proizvodi_u_korpi
    
        parts = selected_product.split(" - Cena: ")
        naziv = parts[0]
        cena = float(parts[1])
        kolicina = 1  # Postavljamo količinu na 1 za sada
        ukupan_iznos = cena * kolicina
        proizvodi_u_korpi.append((naziv, cena, kolicina, ukupan_iznos))

    def ukloni_iz_korpe():
        selected_item = korpa_listbox.curselection()
        index = selected_item[0] if selected_item else None
        
        if index is not None:
            korpa_listbox.delete(index)
            # Ukloni izabrani proizvod iz liste proizvodi_u_korpi
            proizvod = proizvodi_u_korpi.pop(index)
            naziv, cena, kolicina, ukupan_iznos = proizvod
            
            # Ažuriranje stanja proizvoda u SQL bazi podataka
            cursor = proizvodi.conn.cursor()
            cursor.execute("UPDATE proizvodi SET stanje = stanje + %s WHERE naziv = %s", (kolicina, naziv))
            proizvodi.conn.commit()
            cursor.close()


    def trazi_proizvod():
        result = proizvodi.trazi_proizvod(search_entry.get())
        search_label_text.configure(text=result)

    def prikazi_racun():
        racun_prozor = tk.Toplevel(kupac_prozor)
        racun_prozor.title("Račun")
        total = 0
        proizvodi_u_korpi = []

        for item in korpa_listbox.get(0, tk.END):
            parts = item.split(" - Cena: ")
            naziv = parts[0]
            
            if len(parts) == 2:
                cena_kolicina = parts[1]
                cena_parts = cena_kolicina.split(" - Količina: ")
                cena = float(cena_parts[0])
                
                if len(cena_parts) == 2:
                    kolicina = int(cena_parts[1])
                else:
                    kolicina = 1  # Postavljamo količinu na 1 ako nije eksplicitno navedena
            else:
                cena = 0
                kolicina = 1  # Postavljamo količinu na 1 ako nema informacija o ceni
            
            ukupan_iznos = cena * kolicina

            total += ukupan_iznos
            proizvodi_u_korpi.append((naziv, cena, kolicina, ukupan_iznos))


        # Prikaz proizvoda u korpi
        proizvod_frame = tk.Frame(racun_prozor)
        proizvod_frame.pack(pady=10)

        proizvod_label = tk.Label(proizvod_frame, text="Proizvodi u korpi:")
        proizvod_label.pack()

        proizvod_listbox = tk.Listbox(proizvod_frame)
        proizvod_listbox.pack()

        for proizvod in proizvodi_u_korpi:
            naziv, cena, kolicina, ukupan_iznos = proizvod
            item = f"{naziv} - Cena: {cena:.2f} - Količina: {kolicina} - Ukupan iznos: {ukupan_iznos}"
            proizvod_listbox.insert(tk.END, item)

        # Prikaži ukupan iznos
        ukupan_iznos_label = tk.Label(racun_prozor, text=f"Ukupan iznos: {total:.2f} RSD")
        ukupan_iznos_label.pack(pady=10)

        #Stampanje računa
        racun_button_1 = tk.Button(racun_prozor, text="Štampaj račun", command=stampaj_racun)
        racun_button_1.pack(pady=5)

    def stampaj_racun():
        # Kreiranje DataFrame-a sa proizvodima u korpi
        data = []
        total = 0  # Ukupan iznos računa
        kolicine_za_azuriranje = {}  # Rečnik za čuvanje količina za ažuriranje stanja

        for proizvod in proizvodi_u_korpi:
            naziv, cena, kolicina, ukupan_iznos = proizvod

            # Provera da li se proizvod već nalazi u listi data
            found = False
            for i, d in enumerate(data):
                if d['Naziv'] == naziv:
                    found = True
                    data[i]['Količina'] += kolicina
                    data[i]['Ukupan iznos'] += ukupan_iznos
                    break

            if not found:
                data.append({'Naziv': naziv, 'Cena': cena, 'Količina': kolicina, 'Ukupan iznos': ukupan_iznos})

            total += ukupan_iznos

            # Dodavanje količine za ažuriranje stanja u rečnik
            if naziv in kolicine_za_azuriranje:
                kolicine_za_azuriranje[naziv] += kolicina
            else:
                kolicine_za_azuriranje[naziv] = kolicina

        # Ažuriranje stanja proizvoda u SQL bazi podataka
        cursor = proizvodi.conn.cursor()
        for naziv, kolicina in kolicine_za_azuriranje.items():
            cursor.execute("UPDATE proizvodi SET stanje = stanje - %s WHERE naziv = %s", (kolicina, naziv))
        proizvodi.conn.commit()
        cursor.close()

        df = pd.DataFrame(data)

        # Snimanje DataFrame-a u Excel fajl
        file_path = "racun.xlsx"
        df.to_excel(file_path, index=False)

        # Dodavanje ukupnog iznosa na listu proizvoda u korpi
        df.loc[len(df)] = ['Ukupan iznos:', '', '', total]

        # Snimanje ažuriranog DataFrame-a u Excel fajl
        df.to_excel(file_path, index=False)

        # Prikaz prozora za potvrdu
        pyautogui.alert("Račun je uspešno izvezen u Excel fajl.")

    




    # Search Section
    search_frame = tk.Frame(kupac_prozor)
    search_frame.pack(pady=10)

    search_label = tk.Label(search_frame, text="Pretraga:")
    search_label.pack(side=tk.LEFT)

    search_entry = tk.Entry(search_frame)
    search_entry.pack(side=tk.LEFT)

    search_button = tk.Button(search_frame, text="Traži",
                              command=trazi_proizvod)
    search_button.pack(side=tk.LEFT)
    
    search_label_text = tk.Label(search_frame, text='')
    search_label_text.pack()

    # Korpa
    korpa_frame = tk.Frame(kupac_prozor)
    korpa_frame.pack(pady=10)

    korpa_label = tk.Label(korpa_frame, text="Korpa:")
    korpa_label.pack()

    korpa_listbox = tk.Listbox(korpa_frame)
    korpa_listbox.pack()

    ukloni_button = tk.Button(kupac_prozor, text="Ukloni iz korpe", command=ukloni_iz_korpe)
    ukloni_button.pack(pady=5)

    racun_button = tk.Button(kupac_prozor, text="Račun", command=prikazi_racun)
    racun_button.pack(pady=5)

    dodaj_button = tk.Button(kupac_prozor, text="Dodaj u korpu", command=dodaj_u_korpu)
    dodaj_button.pack(pady=5)



    # Proizvodi
    proizvod_frame = tk.Frame(kupac_prozor)
    proizvod_frame.pack(pady=10)

    proizvod_label = tk.Label(proizvod_frame, text="Proizvodi:")
    proizvod_label.pack()

    proizvod_listbox = tk.Listbox(proizvod_frame)
    proizvod_listbox.pack()

    proizvodi_df = ucitaj_proizvode()
    for index, row in proizvodi_df.iterrows():
        naziv = row['naziv']
        cena = row['cena']
        item = f"{naziv} - Cena: {cena}"
        proizvod_listbox.insert(tk.END, item)



def admin_prozor():
    admin_prozor = tk.Toplevel(root)
    admin_prozor.title("Administrator")
    izvestaji_prozor = None  # Globalna promenljiva za izvestaji_prozor

    def proveri_prijavu():
        uneto_ime = entry_ime.get()
        uneta_sifra = entry_sifra.get()

        df_administratori = administrator.ucitaj_administratora()

        def prikazi_izvestaj(izvestaj):
                # Implementirajte logiku za prikaz izveštaja u odabranom formatu (Excel, JSON, CSV)
                if izvestaj == "Excel":             
                  
                    # Učitaj sve račune iz Excel fajla
                    wb = openpyxl.load_workbook('racun.xlsx')
                    sheet = wb.active

                    # Kreiraj novi Excel fajl za izveštaj
                    izvestaj_wb = openpyxl.Workbook()
                    izvestaj_sheet = izvestaj_wb.active

                    # Kopiraj podatke o računima u izveštajni Excel fajl
                    row_index = 1
                    for row in sheet.iter_rows(values_only=True):
                        if row[0] == 'Ukupan iznos:':
                            # Kraj izveštaja
                            break
                        izvestaj_sheet.cell(row=row_index, column=1).value = row[0]
                        izvestaj_sheet.cell(row=row_index, column=2).value = row[1]
                        izvestaj_sheet.cell(row=row_index, column=3).value = row[2]
                        izvestaj_sheet.cell(row=row_index, column=4).value = row[3]
                        row_index += 1

                    # Sačuvaj izveštajni Excel fajl
                    izvestaj_wb.save('izvestaj.xlsx')

                    # Zatvori obe Excel datoteke
                    wb.close()
                    izvestaj_wb.close()

                    print("Izveštaj je generisan i sačuvan u izvestaj.xlsx.")

                elif izvestaj == "JSON":

                    df = pd.read_excel('racun.xlsx')
                    
                    izvestaj = df.to_json(orient='records')

                    # Sačuvaj izveštaj u JSON fajl
                    with open('izvestaj.json', 'w') as f:
                        f.write(izvestaj)

                    print("Izveštaj je generisan i sačuvan u izvestaj.json.")

                elif izvestaj == "CSV":
                    df = pd.read_excel('racun.xlsx')
                    
                    # Sačuvaj izveštaj u CSV fajl
                    df.to_csv('izvestaj.csv', index=False)

                    print("Izveštaj je generisan i sačuvan u izvestaj.csv.")

        def prikazi_grafik():
            ## Učitaj podatke iz Excel fajla
            df = pd.read_excel('izvestaj.xlsx')

            # Izdvoj kolone za proizvode i prodaju
            proizvodi = df['Naziv']
            prodaja = df['Ukupan iznos']

            # Prikazi grafik
            plt.plot(proizvodi, prodaja)
            plt.xlabel('Proizvodi')
            plt.ylabel('Prodaja')
            plt.title('Grafik prodaje')
            plt.show()

        def prikazi_lager():
            nonlocal izvestaji_prozor  # Pristupanje izvestaji_prozor-u
            proizvodi.ucitaj_proizvode()
            lager = proizvodi.ucitaj_proizvode()
            text_lagera = tk.Text(izvestaji_prozor)
            text_lagera.insert(tk.END, f"Finalno stanje lagera:\n{lager}")
            text_lagera.config(state=tk.DISABLED)
            text_lagera.pack()

        def prozor_izvestaji():
            nonlocal izvestaji_prozor  # Pristupanje izvestaji_prozor-u
            izvestaji_prozor = tk.Toplevel(root)
            izvestaji_prozor.title("Statistika prodaje")

            # Padajući meni za izbor formata izveštaja
            izvestaj_var = tk.StringVar(izvestaji_prozor)
            izvestaj_var.set("Izveštaj")  # Početni izbor
            izvestaj_menu = tk.OptionMenu(izvestaji_prozor, izvestaj_var, "Excel", "JSON", "CSV")
            izvestaj_menu.pack()

            # Dugme za prikaz izveštaja
            dugme_izvestaj = tk.Button(izvestaji_prozor, text="Prikaži izveštaj", command=lambda: prikazi_izvestaj(izvestaj_var.get()))
            dugme_izvestaj.pack()

            # Dugme za prikaz grafa
            dugme_grafik = tk.Button(izvestaji_prozor, text="Prikaži grafik", command=prikazi_grafik)
            dugme_grafik.pack()

            # Dugme za prikaz lagera
            dugme_lager = tk.Button(izvestaji_prozor, text="Prikaži lager", command=prikazi_lager)
            dugme_lager.pack()

        if any((df_administratori['ime'] == uneto_ime) & (df_administratori['password'] == uneta_sifra)):
            # Prijavljivanje uspešno
            prozor_izvestaji()
            
        else:
            pyautogui.alert("Prijavljivanje neuspešno!")

        def prikazi_lager():
            proizvodi.ucitaj_proizvode()
            labela_lagera = tk.Label(izvestaji_prozor,text=f"Finalno stanje lagera: {proizvodi.ucitaj_proizvode()}")
            labela_lagera.pack()

    def prijava():
    # Provera korisničkih podataka sa bazom podataka
        proveri_prijavu()


    # Unosno polje za ime
    label_ime = tk.Label(admin_prozor, text="Ime:")
    label_ime.pack()
    entry_ime = tk.Entry(admin_prozor)
    entry_ime.pack()

    # Unosno polje za šifru
    label_sifra = tk.Label(admin_prozor, text="Šifra:")
    label_sifra.pack()
    entry_sifra = tk.Entry(admin_prozor, show="*")
    entry_sifra.pack()

    # Dugme za prijavu
    dugme_prijava = tk.Button(admin_prozor, text="Prijavi se", command=prijava)
    dugme_prijava.pack()



root = tk.Tk()
root.title("Knjižara Bookstore!")

welcome_label = tk.Label(root, text="Dobrodošli!")
welcome_label.pack(pady=10)

customer_button = tk.Button(root, text="Kupac", command=kupac_prozor)
customer_button.pack(pady=5)

admin_button = tk.Button(root, text="Administrator", command=admin_prozor)
admin_button.pack(pady=5)


img = ImageTk.PhotoImage(Image.open("download.jpg"))
image_label = tk.Label(root, image=img)
image_label.pack(pady=5)

root.mainloop()